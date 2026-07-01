"""XLSX text extractor using openpyxl."""

import os

from . import TextExtractor, ExtractedText


class XlsxExtractor(TextExtractor):
	"""Extractor for XLSX / XLS Excel files."""

	def extract(self, file_path: str) -> ExtractedText:
		"""Extract text from an Excel workbook."""
		try:
			from openpyxl import load_workbook
		except ImportError:
			raise ImportError(
				"XLSX extraction requires 'openpyxl'. "
				"Install with: pip install openpyxl"
			)

		wb = load_workbook(file_path, data_only=True, read_only=True)
		sheet_count = len(wb.sheetnames)
		text_parts = []

		for sheet in wb.worksheets:
			sheet_lines = [f"## Sheet: {sheet.title}"]
			for row in sheet.iter_rows(values_only=True):
				row_text = "\t".join(
					str(cell) if cell is not None else "" for cell in row
				).strip()
				if row_text:
					sheet_lines.append(row_text)
			if len(sheet_lines) > 1:
				text_parts.append("\n".join(sheet_lines))

		wb.close()

		return ExtractedText(
			text="\n\n".join(text_parts),
			title=os.path.basename(file_path),
			metadata={"file_type": "xlsx", "sheets": sheet_count},
		)
